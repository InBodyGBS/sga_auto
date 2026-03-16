"""
Excel 파일 생성 모듈
- Pivot / SG&A / Analysis / Analysis_SG&A / Sheet_total 시트 생성
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

from config import (
    REV_ACCOUNT_ORDER,
    ENTITY_ORDER,
    SGA_TARGET_ACCOUNTS,
    SGA_COMPONENTS,
    CALCULATED_ACCOUNTS,
    LABOR_ACCOUNTS,
    MISC_SGA_ACCOUNTS,
)

# ===========================================================
# 스타일 상수
# ===========================================================
STYLE = {
    "header":      PatternFill("solid", fgColor="4472C4"),
    "subheader":   PatternFill("solid", fgColor="D9E1F2"),
    "period_acc":  PatternFill("solid", fgColor="4472C4"),  # 누적 헤더
    "period_qtd":  PatternFill("solid", fgColor="70AD47"),  # 분기별 헤더
    "section":     PatternFill("solid", fgColor="ED7D31"),  # 법인 구분 헤더
    "calc":        PatternFill("solid", fgColor="FFF2CC"),  # 계산 계정
    "increase":    PatternFill("solid", fgColor="C6EFCE"),  # 증가
    "decrease":    PatternFill("solid", fgColor="FFC7CE"),  # 감소
}
FONT_WHITE_BOLD  = Font(bold=True, color="FFFFFF")
FONT_BOLD_LG     = Font(bold=True, size=12, color="FFFFFF")
FONT_BOLD        = Font(bold=True)
CENTER           = Alignment(horizontal="center", vertical="center")
RIGHT            = Alignment(horizontal="right", vertical="center")


def _set(cell, value=None, fill=None, font=None, align=None, fmt=None):
    """셀 속성 일괄 설정 헬퍼"""
    if value is not None:
        cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt


# ===========================================================
# 공통 유틸
# ===========================================================

def sort_rev_accounts(accounts: list) -> list:
    ordered = [a for a in REV_ACCOUNT_ORDER if a in accounts]
    remaining = sorted(a for a in accounts if a not in REV_ACCOUNT_ORDER)
    return ordered + remaining


def sort_entities(entities: list) -> list:
    ordered = [e for e in ENTITY_ORDER if e in entities]
    remaining = sorted(e for e in entities if e not in ENTITY_ORDER)
    return ordered + remaining


def _add_calculated_to_dict(data: dict) -> dict:
    """인건비 / 기타 / 매출총이익 / 판관비 / 영업이익 계산 후 dict에 주입"""
    data["인건비"]     = sum(data.get(acc, 0.0) for acc in LABOR_ACCOUNTS)
    data["기타"]       = sum(data.get(acc, 0.0) for acc in MISC_SGA_ACCOUNTS)
    data["매출총이익"] = data.get("매출", 0.0) - data.get("매출원가", 0.0)
    data["판관비"]     = sum(data.get(acc, 0.0) for acc in SGA_COMPONENTS)
    data["영업이익"]   = data["매출총이익"] - data["판관비"]
    return data


def parse_period(period_str: str):
    """'2025_Q4' → (2025, 4)"""
    import re
    if not period_str or not isinstance(period_str, str):
        return None, None
    if period_str in ("전체", "합계", "Total"):
        return None, None
    m = re.match(r'^(\d{4})_Q([1-4])$', period_str)
    if m:
        return int(m.group(1)), int(m.group(2))
    try:
        return int(period_str[:4]), int(period_str[4])
    except (ValueError, IndexError):
        return None, None


def fmt_period(period_str: str) -> str:
    y, q = parse_period(period_str)
    return f"{y}_Q{q}" if y else period_str


def get_prev_quarter(period_str: str) -> str | None:
    y, q = parse_period(period_str)
    if y is None:
        return None
    return f"{y-1}_Q4" if q == 1 else f"{y}_Q{q-1}"


def get_yoy_quarter(period_str: str) -> str | None:
    y, q = parse_period(period_str)
    return f"{y-1}_Q{q}" if y else None


def _get_pivot_section(ws, section_keyword: str, entity_name: str) -> dict:
    """Pivot 시트에서 특정 섹션 + 법인 데이터 추출 → {account: value}"""
    data = {}
    max_row = ws.max_row
    row = 1
    while row <= max_row:
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and isinstance(cell_val, str) and section_keyword in cell_val:
            row += 1
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value == entity_name:
                    target_col = col
                    break
            if not target_col:
                row += 1
                continue
            row += 1
            while row <= max_row:
                rev_account = ws.cell(row=row, column=1).value
                if not rev_account or (isinstance(rev_account, str) and
                        ("누적" in rev_account or "분기별" in rev_account)):
                    break
                val = ws.cell(row=row, column=target_col).value
                try:
                    data[rev_account] = float(val) if val else 0.0
                except (ValueError, TypeError):
                    data[rev_account] = 0.0
                row += 1
            break
        row += 1
    return data


# ===========================================================
# Step 1: Pivot 시트
# ===========================================================

def _write_pivot_ws(ws, pivot_df, label_suffix: str = ""):
    """pivot_df → 워크시트에 누적/분기별 섹션 작성"""
    periods, entities_by_period = [], {}
    for col in pivot_df.columns:
        if col == "Rev_Account" or (isinstance(col, tuple) and col[0] == "Rev_Account"):
            continue
        period, entity = col
        if period not in periods:
            periods.append(period)
            entities_by_period[period] = []
        if entity not in entities_by_period[period]:
            entities_by_period[period].append(entity)

    for period in periods:
        entities_by_period[period] = sort_entities(entities_by_period[period])

    rev_accounts = sort_rev_accounts(pivot_df["Rev_Account"].tolist())
    current_row = 1
    prev_data: dict | None = None

    for idx, period in enumerate(periods):
        if period in ("전체", "Rev_Account"):
            continue
        entities = entities_by_period[period]
        n_cols = len(entities) + 1  # account col + entity cols

        # ── 누적 헤더 ──
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=n_cols)
        _set(ws.cell(current_row, 1), f"{period} (누적){label_suffix}",
             STYLE["period_acc"], FONT_BOLD_LG, CENTER)
        current_row += 1

        _set(ws.cell(current_row, 1), "Rev_Account",
             STYLE["header"], FONT_WHITE_BOLD, CENTER)
        for ci, ent in enumerate(entities, 2):
            _set(ws.cell(current_row, ci), ent,
                 STYLE["subheader"], FONT_BOLD, CENTER)
        current_row += 1

        cur_data: dict[str, dict] = {}
        for account in rev_accounts:
            ws.cell(current_row, 1).value = account
            row_src = pivot_df[pivot_df["Rev_Account"] == account]
            for ci, ent in enumerate(entities, 2):
                key = (period, ent)
                try:
                    val = float(row_src.iloc[0][key]) if (
                        not row_src.empty and key in pivot_df.columns) else 0.0
                except (ValueError, TypeError):
                    # 숫자로 변환 불가능한 경우 (헤더나 문자열)
                    val = 0
                ws.cell(current_row, ci).value = val
                cur_data.setdefault(account, {})[ent] = val
            current_row += 1

        # ── 분기별 섹션 (2Q 이후부터) ──
        if idx > 0 and prev_data is not None:
            y, q = parse_period(period)
            if y is None:
                prev_data = cur_data
                continue
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=n_cols)
            _set(ws.cell(current_row, 1), f"{period} (분기별){label_suffix}",
                 STYLE["period_qtd"], FONT_BOLD_LG, CENTER)
            current_row += 1

            _set(ws.cell(current_row, 1), "Rev_Account",
                 STYLE["header"], FONT_WHITE_BOLD, CENTER)
            for ci, ent in enumerate(entities, 2):
                _set(ws.cell(current_row, ci), ent,
                     STYLE["subheader"], FONT_BOLD, CENTER)
            current_row += 1

            for account in rev_accounts:
                ws.cell(current_row, 1).value = account
                for ci, ent in enumerate(entities, 2):
                    cur_val = cur_data.get(account, {}).get(ent, 0.0)
                    if q == 1:
                        qtd = cur_val
                    else:
                        prev_val = prev_data.get(account, {}).get(ent, 0.0)
                        qtd = cur_val - prev_val
                    ws.cell(current_row, ci).value = qtd
                current_row += 1

        prev_data = cur_data
        if idx < len(periods) - 1:
            current_row += 2

    # 열 너비
    ws.column_dimensions["A"].width = 28
    max_ents = max((len(v) for v in entities_by_period.values()), default=1)
    for ci in range(2, max_ents + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def create_pivot_sheet(wb: openpyxl.Workbook, pivot_df, sheet_name: str = "Pivot"):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    suffix = " (FCY)" if "FCY" in sheet_name else ""
    _write_pivot_ws(ws, pivot_df, suffix)
    print(f"  ✓ {sheet_name} 시트 생성 완료")
    return ws


# ===========================================================
# Step 2: SG&A 시트 (Pivot 시트 기반)
# ===========================================================

def _sum_accounts(account_list: list, section_data: dict, col_count: int) -> list:
    """account_list 에 속한 계정 값을 합산"""
    row = [0.0] * col_count
    for comp in account_list:
        vals = section_data.get(comp, [0.0] * (col_count + 1))
        for i in range(col_count):
            try:
                row[i] += float(vals[i + 1] or 0)
            except (TypeError, ValueError):
                pass
    return row


def _compute_calculated_row(account: str, section_data: dict, col_count: int) -> list:
    """매출총이익 / 인건비 / 기타 / 판관비 / 영업이익 계산"""
    if account == "매출총이익":
        rev  = section_data.get("매출",    [0.0] * (col_count + 1))
        cogs = section_data.get("매출원가", [0.0] * (col_count + 1))
        row = [0.0] * col_count
        for i in range(col_count):
            try:
                row[i] = float(rev[i + 1] or 0) - float(cogs[i + 1] or 0)
            except (TypeError, ValueError):
                row[i] = 0.0
        return row
    elif account == "인건비":
        return _sum_accounts(LABOR_ACCOUNTS, section_data, col_count)
    elif account == "기타":
        return _sum_accounts(MISC_SGA_ACCOUNTS, section_data, col_count)
    elif account == "판관비":
        return _sum_accounts(SGA_COMPONENTS, section_data, col_count)
    elif account == "영업이익":
        gp = _compute_calculated_row("매출총이익", section_data, col_count)
        sg = _compute_calculated_row("판관비",    section_data, col_count)
        return [g - s for g, s in zip(gp, sg)]
    return [0.0] * col_count


def create_sga_sheet(wb: openpyxl.Workbook, pivot_sheet_name: str = "Pivot",
                     sga_sheet_name: str = "SG&A"):
    if pivot_sheet_name not in wb.sheetnames:
        print(f"  ❌ {pivot_sheet_name} 시트 없음")
        return
    if sga_sheet_name in wb.sheetnames:
        del wb[sga_sheet_name]

    pivot_ws = wb[pivot_sheet_name]
    sga_ws = wb.create_sheet(sga_sheet_name)

    max_row = pivot_ws.max_row
    max_col = pivot_ws.max_column
    sga_row = 1
    cur_pivot_row = 1
    section_count = 0

    while cur_pivot_row <= max_row:
        header_val = pivot_ws.cell(cur_pivot_row, 1).value
        if not (header_val and isinstance(header_val, str) and
                ("누적" in header_val or "분기별" in header_val)):
            cur_pivot_row += 1
            continue

        section_count += 1
        is_qtd = "분기별" in header_val

        # 섹션 헤더 복사
        sga_ws.merge_cells(start_row=sga_row, start_column=1,
                           end_row=sga_row, end_column=max_col)
        _set(sga_ws.cell(sga_row, 1), header_val,
             STYLE["period_qtd"] if is_qtd else STYLE["period_acc"],
             FONT_BOLD_LG, CENTER)
        sga_row += 1
        cur_pivot_row += 1

        # 컬럼 헤더 복사
        for col in range(1, max_col + 1):
            src = pivot_ws.cell(cur_pivot_row, col)
            dst = sga_ws.cell(sga_row, col)
            _set(dst, src.value,
                 STYLE["header"] if col == 1 else STYLE["subheader"],
                 FONT_WHITE_BOLD if col == 1 else FONT_BOLD, CENTER)
        sga_row += 1
        cur_pivot_row += 1

        # 데이터 수집
        section_data: dict[str, list] = {}
        while cur_pivot_row <= max_row:
            acct = pivot_ws.cell(cur_pivot_row, 1).value
            if not acct or (isinstance(acct, str) and
                            ("누적" in acct or "분기별" in acct)):
                break
            row_vals = [acct] + [
                pivot_ws.cell(cur_pivot_row, c).value
                for c in range(2, max_col + 1)
            ]
            section_data[acct] = row_vals
            cur_pivot_row += 1

        # 대상 계정 출력
        data_col_count = max_col - 1
        for acct in SGA_TARGET_ACCOUNTS:
            if acct in CALCULATED_ACCOUNTS:
                calc_vals = _compute_calculated_row(acct, section_data, data_col_count)
                sga_ws.cell(sga_row, 1).value = acct
                _set(sga_ws.cell(sga_row, 1), fill=STYLE["calc"], font=FONT_BOLD)
                for ci, val in enumerate(calc_vals, 2):
                    sga_ws.cell(sga_row, ci).value = val
                    sga_ws.cell(sga_row, ci).fill = STYLE["calc"]
            elif acct in section_data:
                row_vals = section_data[acct]
                for ci, val in enumerate(row_vals):
                    sga_ws.cell(sga_row, ci + 1).value = val
            sga_row += 1

        sga_row += 1  # 섹션 간 빈 줄

    sga_ws.column_dimensions["A"].width = 28
    for ci in range(2, max_col + 1):
        sga_ws.column_dimensions[get_column_letter(ci)].width = 18

    print(f"  ✓ {sga_sheet_name} 시트 생성 완료 ({section_count}개 섹션)")


# ===========================================================
# Step 3: Analysis / Analysis_SG&A 시트
# ===========================================================

def _write_comparison_table(ws, start_row: int, entity_label: str,
                             qoq_b: dict, qoq_c: dict,
                             yoyq_b: dict, yoyq_c: dict,
                             yoyy_b: dict, yoyy_c: dict,
                             prev_disp: str, tgt_disp: str, yoy_disp: str,
                             account_list: list) -> int:
    """QoQ | YoY(Q) | YoY(Y) 비교 테이블 1개 작성, 다음 시작 row 반환"""
    row = start_row

    # 법인 헤더
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    _set(ws.cell(row, 1), entity_label, STYLE["section"], FONT_BOLD_LG, CENTER)
    row += 1

    # 컬럼 헤더
    headers = [
        "Rev_Account",
        prev_disp, tgt_disp, "증감액", "증감률(%)",
        yoy_disp, tgt_disp, "증감액", "증감률(%)",
        f"{yoy_disp}(누적)", f"{tgt_disp}(누적)", "증감액", "증감률(%)",
    ]
    for ci, h in enumerate(headers, 1):
        _set(ws.cell(row, ci), h, STYLE["header"], FONT_WHITE_BOLD, CENTER)
    row += 1

    def _write_delta(ws_row, col_start, base: dict, comp: dict, acct: str):
        b = base.get(acct, 0.0)
        c = comp.get(acct, 0.0)
        delta = c - b
        rate = (delta / b * 100) if b != 0 else (0.0 if delta == 0 else float("inf"))
        ws.cell(ws_row, col_start).value = b
        ws.cell(ws_row, col_start + 1).value = c
        delta_cell = ws.cell(ws_row, col_start + 2)
        delta_cell.value = delta
        delta_cell.fill = STYLE["increase"] if delta > 0 else (
            STYLE["decrease"] if delta < 0 else PatternFill())
        rate_cell = ws.cell(ws_row, col_start + 3)
        if rate == float("inf"):
            rate_cell.value = "N/A"
        else:
            rate_cell.value = round(rate, 2)
            rate_cell.number_format = "0.00"
            rate_cell.fill = STYLE["increase"] if rate > 0 else (
                STYLE["decrease"] if rate < 0 else PatternFill())

    for acct in account_list:
        ws.cell(row, 1).value = acct
        if acct in CALCULATED_ACCOUNTS:
            ws.cell(row, 1).fill = STYLE["calc"]
            ws.cell(row, 1).font = FONT_BOLD
        _write_delta(row, 2,  qoq_b,  qoq_c,  acct)
        _write_delta(row, 6,  yoyq_b, yoyq_c, acct)
        _write_delta(row, 10, yoyy_b, yoyy_c, acct)
        row += 1

    return row + 1  # 빈 줄 포함


def _create_analysis_ws(wb: openpyxl.Workbook, pivot_sheet_name: str,
                         analysis_sheet_name: str, target_period: str,
                         account_filter: list | None = None):
    """
    Analysis 계열 시트 공통 생성기
    account_filter=None 이면 모든 계정, 리스트 지정 시 해당 계정만
    """
    if pivot_sheet_name not in wb.sheetnames:
        print(f"  ❌ {pivot_sheet_name} 시트 없음")
        return
    if analysis_sheet_name in wb.sheetnames:
        del wb[analysis_sheet_name]

    pivot_ws = wb[pivot_sheet_name]

    prev_q   = target_period and get_prev_quarter(target_period)
    yoy_q    = target_period and get_yoy_quarter(target_period)
    tgt_disp = fmt_period(target_period)
    prev_disp = fmt_period(prev_q) if prev_q else "N/A"
    yoy_disp  = fmt_period(yoy_q)  if yoy_q  else "N/A"

    # 섹션 키 구성
    tgt_qtd_key  = f"{target_period} (분기별)"
    tgt_acc_key  = f"{target_period} (누적)"
    prev_qtd_key = f"{prev_q} (분기별)" if prev_q else None
    yoyq_qtd_key = f"{yoy_q} (분기별)" if yoy_q else None
    yoyy_acc_key = f"{yoy_q} (누적)"   if yoy_q else None

    # 발견된 법인 목록
    entities = set()
    for row in range(1, pivot_ws.max_row + 1):
        v = pivot_ws.cell(row, 1).value
        if v and isinstance(v, str) and tgt_qtd_key in v:
            row += 1
            for col in range(2, pivot_ws.max_column + 1):
                ent = pivot_ws.cell(row, col).value
                if ent and ent != "합계":
                    entities.add(ent)
            break

    entities_sorted = sort_entities(list(entities))

    ws = wb.create_sheet(analysis_sheet_name)
    current_row = 1

    # 분석 분기 정보 헤더
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=13)
    _set(ws.cell(current_row, 1),
         f"【분석 기준: {tgt_disp}】  QoQ: {prev_disp}→{tgt_disp}  |  "
         f"YoY(Q): {yoy_disp}→{tgt_disp}  |  "
         f"YoY(Y): {yoy_disp}(누적)→{tgt_disp}(누적)",
         STYLE["period_acc"], FONT_BOLD_LG, CENTER)
    current_row += 2

    # 전체 합계 포함 순서
    all_entities = ["합계"] + entities_sorted

    for entity in all_entities:
        qoq_b  = _get_pivot_section(pivot_ws, tgt_qtd_key,  entity) if prev_qtd_key  else {}
        if prev_qtd_key:
            qoq_b = _get_pivot_section(pivot_ws, prev_qtd_key, entity)
        qoq_c  = _get_pivot_section(pivot_ws, tgt_qtd_key,  entity)
        yoyq_b = _get_pivot_section(pivot_ws, yoyq_qtd_key, entity) if yoyq_qtd_key else {}
        yoyq_c = _get_pivot_section(pivot_ws, tgt_qtd_key,  entity)
        yoyy_b = _get_pivot_section(pivot_ws, yoyy_acc_key, entity) if yoyy_acc_key else {}
        yoyy_c = _get_pivot_section(pivot_ws, tgt_acc_key,  entity)

        # 인건비 / 기타 / 매출총이익 / 판관비 / 영업이익 계산값 주입
        for d in (qoq_b, qoq_c, yoyq_b, yoyq_c, yoyy_b, yoyy_c):
            _add_calculated_to_dict(d)

        # 계정 목록 결정
        if account_filter:
            all_accts_raw = account_filter
        else:
            all_accts_raw = list(REV_ACCOUNT_ORDER)

        label = "전체 합계" if entity == "합계" else entity
        current_row = _write_comparison_table(
            ws, current_row, label,
            qoq_b, qoq_c, yoyq_b, yoyq_c, yoyy_b, yoyy_c,
            prev_disp, tgt_disp, yoy_disp,
            all_accts_raw,
        )

    # 열 너비
    ws.column_dimensions["A"].width = 28
    for ci in range(2, 14):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    print(f"  ✓ {analysis_sheet_name} 시트 생성 완료 ({len(all_entities)}개 법인)")


def create_analysis_sheet(wb, pivot_sheet: str, sheet_name: str, target_period: str):
    _create_analysis_ws(wb, pivot_sheet, sheet_name, target_period, account_filter=None)


def create_analysis_sga_sheet(wb, pivot_sheet: str, sheet_name: str, target_period: str):
    _create_analysis_ws(wb, pivot_sheet, sheet_name, target_period,
                        account_filter=SGA_TARGET_ACCOUNTS)


# ===========================================================
# Step 4: Sheet_total PL (법인 가로 배치 요약)
# ===========================================================

def create_sheet_total(wb: openpyxl.Workbook, pivot_sheet_name: str,
                       sheet_name: str, target_period: str,
                       account_filter: list | None = None):
    """
    법인을 열로, 계정을 행으로 배치하는 요약 시트 (QoQ / YoY(Q) / YoY(Y) 각 1개 테이블)
    """
    if pivot_sheet_name not in wb.sheetnames:
        print(f"  ❌ {pivot_sheet_name} 시트 없음")
        return
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    pivot_ws = wb[pivot_sheet_name]

    prev_q  = get_prev_quarter(target_period)
    yoy_q   = get_yoy_quarter(target_period)
    tgt_d   = fmt_period(target_period)
    prev_d  = fmt_period(prev_q)  if prev_q else "N/A"
    yoy_d   = fmt_period(yoy_q)   if yoy_q  else "N/A"

    tgt_qtd  = f"{target_period} (분기별)"
    prev_qtd = f"{prev_q} (분기별)" if prev_q else None
    yoyq_qtd = f"{yoy_q} (분기별)"  if yoy_q  else None
    tgt_acc  = f"{target_period} (누적)"
    yoyy_acc = f"{yoy_q} (누적)"    if yoy_q  else None

    # 법인 목록
    entities = set()
    for row in range(1, pivot_ws.max_row + 1):
        v = pivot_ws.cell(row, 1).value
        if v and isinstance(v, str) and tgt_qtd in v:
            row += 1
            for col in range(2, pivot_ws.max_column + 1):
                ent = pivot_ws.cell(row, col).value
                if ent and ent != "합계":
                    entities.add(ent)
            break
    entities_sorted = sort_entities(list(entities)) + ["합계"]

    ws = wb.create_sheet(sheet_name)
    current_row = 1
    n_ents = len(entities_sorted)

    def _write_total_table(start_row, section_label, base_key, comp_key, is_acc=False):
        row = start_row
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_ents * 3 + 1)
        _set(ws.cell(row, 1), section_label,
             STYLE["period_acc"], FONT_BOLD_LG, CENTER)
        row += 1

        # 법인 헤더 (3열씩: 기준, 비교, 증감률)
        ws.cell(row, 1).value = "Rev_Account"
        ws.cell(row, 1).fill = STYLE["header"]
        ws.cell(row, 1).font = FONT_WHITE_BOLD
        ws.cell(row, 1).alignment = CENTER
        for ei, ent in enumerate(entities_sorted):
            c = 2 + ei * 3
            ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + 2)
            _set(ws.cell(row, c), ent, STYLE["subheader"], FONT_BOLD, CENTER)
        row += 1

        # 기준/비교/증감률 서브헤더
        ws.cell(row, 1).value = ""
        for ei in range(n_ents):
            c = 2 + ei * 3
            b_label = f"{fmt_period(base_key.split(' ')[0])}{'(누적)' if is_acc else ''}"
            c_label = f"{tgt_d}{'(누적)' if is_acc else ''}"
            ws.cell(row, c).value = b_label
            ws.cell(row, c + 1).value = c_label
            ws.cell(row, c + 2).value = "증감률(%)"
            for ci in [c, c + 1, c + 2]:
                ws.cell(row, ci).fill = STYLE["subheader"]
                ws.cell(row, ci).font = FONT_BOLD
                ws.cell(row, ci).alignment = CENTER
        row += 1

        # 계정 데이터
        accts_raw = account_filter if account_filter else SGA_TARGET_ACCOUNTS
        # 엔티티별 데이터 사전 로드 + 계산 계정(인건비/기타/매출총이익/판관비/영업이익) 주입
        base_by_ent = {ent: _add_calculated_to_dict(
            _get_pivot_section(pivot_ws, base_key, ent) if base_key else {}
        ) for ent in entities_sorted}
        comp_by_ent = {ent: _add_calculated_to_dict(
            _get_pivot_section(pivot_ws, comp_key, ent)
        ) for ent in entities_sorted}

        for acct in accts_raw:
            ws.cell(row, 1).value = acct
            is_calc = acct in CALCULATED_ACCOUNTS
            if is_calc:
                ws.cell(row, 1).fill = STYLE["calc"]
                ws.cell(row, 1).font = FONT_BOLD
            for ei, ent in enumerate(entities_sorted):
                b     = base_by_ent[ent].get(acct, 0.0)
                c_val = comp_by_ent[ent].get(acct, 0.0)
                delta = c_val - b
                rate  = (delta / b * 100) if b != 0 else (0.0 if delta == 0 else float("inf"))
                col   = 2 + ei * 3
                ws.cell(row, col).value     = b
                ws.cell(row, col + 1).value = c_val
                if is_calc:
                    ws.cell(row, col).fill     = STYLE["calc"]
                    ws.cell(row, col + 1).fill = STYLE["calc"]
                rate_cell = ws.cell(row, col + 2)
                if rate == float("inf"):
                    rate_cell.value = "N/A"
                else:
                    rate_cell.value = round(rate, 2)
                    rate_cell.number_format = "0.00"
                    rate_cell.fill = STYLE["increase"] if rate > 0 else (
                        STYLE["decrease"] if rate < 0 else PatternFill())
            row += 1

        return row + 2

    current_row = _write_total_table(
        current_row,
        f"QoQ: {prev_d} → {tgt_d}",
        prev_qtd or "", tgt_qtd,
    )
    current_row = _write_total_table(
        current_row,
        f"YoY(Q): {yoy_d} → {tgt_d}",
        yoyq_qtd or "", tgt_qtd,
    )
    current_row = _write_total_table(
        current_row,
        f"YoY(Y): {yoy_d}(누적) → {tgt_d}(누적)",
        yoyy_acc or "", tgt_acc,
        is_acc=True,
    )

    ws.column_dimensions["A"].width = 28
    for ci in range(2, n_ents * 3 + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    print(f"  ✓ {sheet_name} 시트 생성 완료")


# ===========================================================
# 전체 Excel 파일 생성
# ===========================================================

def save_workbook(wb: openpyxl.Workbook, output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ 저장 완료: {output_path}")
