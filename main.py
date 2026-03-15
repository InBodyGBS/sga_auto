"""
InBody SG&A 분기별 분석 자동화
============================================================
실행 방법:
  - run.bat 더블클릭  (권장)
  - 또는: python main.py

매 분기 작업 순서:
  1. data/raw/ 아래에 분기 폴더 생성  예) 2025_Q4
  2. 해당 폴더에 법인 Excel 파일 복사
  3. config.py 의 TARGET_PERIOD, EXCHANGE_RATES 업데이트
  4. run.bat 실행 → output/InBody_SGA_Analysis.xlsx 생성
============================================================
"""

import sys
import openpyxl

import config
from reader import sync_all_quarters, load_all_pl_data, load_exchange_rates
from exporter import (
    create_pivot_sheet,
    create_sga_sheet,
    create_analysis_sheet,
    create_analysis_sga_sheet,
    create_sheet_total,
    save_workbook,
)


def check_exchange_rates(df):
    """환율 누락 경고"""
    if "Exchange_Rate" not in df.columns:
        return
    missing = df[(df["Currency"] != "KRW") &
                 ((df["Exchange_Rate"].isna()) | (df["Exchange_Rate"] == 0))]
    if missing.empty:
        print("  ✓ 모든 외화 환율 정상")
        return
    print(f"\n  ⚠️  환율 누락 {len(missing)}건:")
    for (entity, currency), grp in missing.groupby(["Entity_Short", "Currency"]):
        print(f"    - {entity} ({currency}): {len(grp)}행")
    print("  → config.py 의 EXCHANGE_RATES 를 확인하세요.\n")


def run():
    target = config.TARGET_PERIOD
    print("=" * 60)
    print(f"InBody SG&A 분기별 분석")
    print(f"분석 대상 분기: {target}")
    print("=" * 60)

    # ── 환율 파일 사전 검증 ───────────────────────────────────
    print(f"\n환율 파일 로드 중...")
    try:
        rates = load_exchange_rates()
        print(f"  ✓ {len(rates)}개 분기 환율 로드 완료 (data/exchange_rates.xlsx)")
        if target not in rates:
            print(f"  ⚠️  [{target}] 환율이 exchange_rates.xlsx 에 없습니다.")
            print(f"      파일을 열어 {target} 분기 행을 추가하세요.")
    except FileNotFoundError as e:
        print(f"\n❌ {e}")
        input("\n아무 키나 누르면 종료...")
        sys.exit(1)

    # ── Step 1: data/raw/ 폴더 자동 스캔 + 신규 분기 통합 ───
    print(f"\n[Step 1] 분기 폴더 스캔 및 데이터 통합")
    sync_all_quarters()

    # ── Step 2: 전체 데이터 로드 ─────────────────────────────
    print(f"\n[Step 2] 전체 기간 데이터 로드")
    df = load_all_pl_data()
    if df is None:
        print("\n❌ 처리할 데이터가 없습니다.")
        print("   data/raw/YYYY_Q# 폴더를 만들고 법인 파일을 넣으세요.")
        input("\n아무 키나 누르면 종료...")
        sys.exit(1)

    check_exchange_rates(df)

    # ── Step 3: Pivot 생성 (원화 / 외화) ─────────────────────
    print(f"\n[Step 3] Pivot 테이블 생성")
    from reader import build_pivot
    pivot_krw = build_pivot(df, "Amount(KRW)")
    pivot_fcy = build_pivot(df, "Amount")

    wb = openpyxl.Workbook()
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    create_pivot_sheet(wb, pivot_krw, "Pivot")
    create_pivot_sheet(wb, pivot_fcy, "Pivot(FCY)")

    # ── Step 4: SG&A 시트 ────────────────────────────────────
    print(f"\n[Step 4] SG&A 시트 생성")
    create_sga_sheet(wb, "Pivot",       "SG&A")
    create_sga_sheet(wb, "Pivot(FCY)",  "SG&A(FCY)")

    # ── Step 5: Analysis (전체 PL) ───────────────────────────
    print(f"\n[Step 5] Analysis 시트 생성 (전체 PL)")
    create_analysis_sheet(wb, "Pivot",      "Analysis",      target)
    create_analysis_sheet(wb, "Pivot(FCY)", "Analysis(FCY)", target)

    # ── Step 6: Analysis_SG&A ────────────────────────────────
    print(f"\n[Step 6] Analysis_SG&A 시트 생성")
    create_analysis_sga_sheet(wb, "Pivot",      "Analysis_SG&A",      target)
    create_analysis_sga_sheet(wb, "Pivot(FCY)", "Analysis_SG&A(FCY)", target)

    # ── Step 7: Sheet_total PL ───────────────────────────────
    print(f"\n[Step 7] Sheet_total 시트 생성")
    create_sheet_total(wb, "Pivot",      "Sheet_total PL",      target)
    create_sheet_total(wb, "Pivot(FCY)", "Sheet_total PL(FCY)", target)

    # ── 저장 ─────────────────────────────────────────────────
    save_workbook(wb, config.OUTPUT_FILE)

    print(f"\n생성된 시트:")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        import traceback
        print(f"\n❌ 오류 발생:\n{traceback.format_exc()}")
    finally:
        input("\n아무 키나 누르면 종료...")
