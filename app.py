"""
InBody SG&A 분기별 분석 - Web UI (Streamlit)
실행: run_web.bat 더블클릭
"""

import io
import sys
import contextlib
import traceback
from pathlib import Path

import pandas as pd
import streamlit as st

# ── 프로젝트 루트를 sys.path 에 추가 ─────────────────────────
ROOT = Path(__file__).parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import config
from reader import (
    RAW_FOLDER,
    RATES_FILE,
    discover_raw_folders,
    load_exchange_rates,
    get_exchange_rates,
    folder_to_period,
)

# ===========================================================
# 페이지 설정
# ===========================================================
st.set_page_config(
    page_title="InBody SG&A 분석",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("📊 InBody SG&A 분기별 분석")

# ===========================================================
# 사이드바 - 분기 선택
# ===========================================================
with st.sidebar:
    st.header("⚙️ 설정")

    # data/raw/ 폴더 스캔
    available_folders = discover_raw_folders()

    if not available_folders:
        st.error("data/raw/ 에 분기 폴더가 없습니다.\n\n폴더명 규칙: `YYYY_Q#`")
        st.stop()

    period_options = sorted(available_folders.keys(), reverse=True)

    # config.py 의 TARGET_PERIOD 를 기본값으로
    default_idx = 0
    if config.TARGET_PERIOD in period_options:
        default_idx = period_options.index(config.TARGET_PERIOD)

    target_period = st.selectbox(
        "분석 대상 분기",
        options=period_options,
        index=default_idx,
        help="data/raw/ 폴더에서 자동으로 감지된 분기 목록입니다.",
    )

    st.divider()

    # 데이터 현황
    st.subheader("📂 데이터 현황")
    processed_folder = config.PROCESSED_FOLDER
    for period in sorted(available_folders.keys()):
        from reader import period_to_filename
        is_processed = (processed_folder / period_to_filename(period)).exists()
        folder = available_folders[period]
        file_count = len([f for f in folder.glob("*.xlsx") if not f.name.startswith("~")])
        if is_processed:
            st.success(f"✓ {period} ({file_count}개 파일)")
        else:
            st.warning(f"⏳ {period} ({file_count}개 파일) - 미처리")

# ===========================================================
# 탭 구성
# ===========================================================
tab_rates, tab_run, tab_result = st.tabs([
    "💱 환율 관리",
    "▶ 분석 실행",
    "📥 결과 다운로드",
])

# ===========================================================
# 탭 1: 환율 관리
# ===========================================================
with tab_rates:
    st.subheader("💱 분기별 환율 관리")
    st.caption(f"파일 위치: `{RATES_FILE}`")

    col1, col2 = st.columns([3, 1])

    with col1:
        # 현재 환율 테이블 표시
        if RATES_FILE.exists():
            df_rates = pd.read_excel(RATES_FILE, header=1)
            df_rates.columns = df_rates.columns.str.strip()
            df_rates.columns = [c if not c.startswith("Rate") else "Rate" for c in df_rates.columns]
            df_rates = df_rates.dropna(subset=["Period", "Currency", "Rate"])

            # 편집 가능한 테이블
            edited = st.data_editor(
                df_rates,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Period":   st.column_config.TextColumn("Period", width="small"),
                    "Currency": st.column_config.TextColumn("Currency", width="small"),
                    "Rate":     st.column_config.NumberColumn(
                        "Rate (1외화 = ?원)",
                        format="%.4f",
                        width="medium",
                    ),
                },
                key="rates_editor",
            )

            if st.button("💾 저장", type="primary", key="save_rates"):
                try:
                    # 기존 서식을 유지하며 데이터만 덮어쓰기
                    import openpyxl
                    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

                    wb = openpyxl.load_workbook(RATES_FILE)
                    ws = wb.active

                    thin = Side(style="thin", color="AAAAAA")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    center = Alignment(horizontal="center", vertical="center")
                    right_align = Alignment(horizontal="right", vertical="center")

                    # 기존 데이터 행 삭제 후 재작성 (헤더 2행 유지)
                    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None

                    row_idx = 3
                    for _, row_data in edited.dropna(subset=["Period", "Currency", "Rate"]).iterrows():
                        ws.cell(row_idx, 1, str(row_data["Period"]).strip()).alignment = center
                        ws.cell(row_idx, 2, str(row_data["Currency"]).strip()).alignment = center
                        rate_cell = ws.cell(row_idx, 3, float(row_data["Rate"]))
                        rate_cell.alignment = right_align
                        rate_cell.number_format = "#,##0.0000"
                        for ci in range(1, 4):
                            ws.cell(row_idx, ci).border = border
                        row_idx += 1

                    wb.save(RATES_FILE)

                    # 캐시 초기화
                    import reader as _reader
                    _reader._EXCHANGE_RATES = None

                    st.success("✅ 저장 완료!")
                    st.rerun()
                except Exception as e:
                    st.error(f"저장 실패: {e}")
        else:
            st.error("exchange_rates.xlsx 파일이 없습니다.")

    with col2:
        st.markdown("**새 분기 추가 방법**")
        st.markdown("""
1. 왼쪽 테이블 하단의 **+** 버튼 클릭
2. Period: `2026_Q1`
3. Currency: `USD` 등 입력
4. Rate: 환율 입력
5. **저장** 클릭
""")

        st.divider()
        st.markdown("**통화 코드 안내**")
        currency_info = {
            "USD": "미국 달러",
            "JPY": "일본 엔",
            "CNH": "중국 위안",
            "AUD": "호주 달러",
            "MXN": "멕시코 페소",
            "VND": "베트남 동",
            "MYR": "말레이시아 링깃",
            "INR": "인도 루피",
            "EUR": "유로",
            "TRY": "터키 리라",
            "KRW": "한국 원",
        }
        for code, name in currency_info.items():
            st.markdown(f"- `{code}`: {name}")

# ===========================================================
# 탭 2: 분석 실행
# ===========================================================
with tab_run:
    st.subheader(f"▶ 분석 실행: **{target_period}**")

    # 사전 점검
    checks_ok = True

    st.markdown("**실행 전 체크리스트**")

    # 1. 환율 체크
    try:
        rates = load_exchange_rates()
        if target_period in rates:
            st.success(f"✓ 환율 설정 완료 ({target_period}: {len(rates[target_period])}개 통화)")
        else:
            st.error(f"✗ [{target_period}] 환율이 없습니다. → 환율 관리 탭에서 추가하세요.")
            checks_ok = False
    except FileNotFoundError as e:
        st.error(str(e))
        checks_ok = False

    # 2. 원본 파일 체크
    folder = available_folders.get(target_period)
    if folder:
        xlsx_files = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~")]
        if xlsx_files:
            st.success(f"✓ 법인 파일 {len(xlsx_files)}개 확인")
        else:
            st.error(f"✗ {folder} 에 xlsx 파일이 없습니다.")
            checks_ok = False
    else:
        st.error(f"✗ data/raw/{target_period}/ 폴더가 없습니다.")
        checks_ok = False

    st.divider()

    # 실행 버튼
    run_btn = st.button(
        f"🚀  {target_period} 분석 실행",
        type="primary",
        disabled=not checks_ok,
        use_container_width=True,
    )

    if run_btn:
        log_area = st.empty()
        log_lines: list[str] = []

        def update_log(text: str):
            log_lines.append(text)
            log_area.code("\n".join(log_lines), language="")

        with st.spinner("분석 중..."):
            try:
                # stdout 캡처하여 실시간 로그 표시
                import openpyxl
                from reader import sync_all_quarters, load_all_pl_data, build_pivot
                from exporter import (
                    create_pivot_sheet, create_sga_sheet,
                    create_analysis_sheet, create_analysis_sga_sheet,
                    create_sheet_total, save_workbook,
                )

                # config.TARGET_PERIOD 임시 변경
                original_period = config.TARGET_PERIOD
                config.TARGET_PERIOD = target_period

                captured = io.StringIO()
                with contextlib.redirect_stdout(captured):

                    update_log(f"[Step 1] 분기 데이터 통합 ({target_period})...")
                    sync_all_quarters()
                    update_log(captured.getvalue())
                    captured.truncate(0); captured.seek(0)

                    update_log("[Step 2] 데이터 로드...")
                    df = load_all_pl_data()
                    update_log(captured.getvalue())
                    captured.truncate(0); captured.seek(0)

                    if df is None:
                        st.error("데이터 로드 실패")
                        config.TARGET_PERIOD = original_period
                        st.stop()

                    update_log("[Step 3] Pivot 생성...")
                    pivot_krw = build_pivot(df, "Amount(KRW)")
                    pivot_fcy = build_pivot(df, "Amount")

                    wb = openpyxl.Workbook()
                    if "Sheet" in wb.sheetnames:
                        del wb["Sheet"]

                    create_pivot_sheet(wb, pivot_krw, "Pivot")
                    create_pivot_sheet(wb, pivot_fcy, "Pivot(FCY)")
                    update_log(captured.getvalue())
                    captured.truncate(0); captured.seek(0)

                    update_log("[Step 4] SG&A 시트 생성...")
                    create_sga_sheet(wb, "Pivot",      "SG&A")
                    create_sga_sheet(wb, "Pivot(FCY)", "SG&A(FCY)")

                    update_log("[Step 5] Analysis 시트 생성...")
                    create_analysis_sheet(wb, "Pivot",      "Analysis",      target_period)
                    create_analysis_sheet(wb, "Pivot(FCY)", "Analysis(FCY)", target_period)

                    update_log("[Step 6] Analysis_SG&A 시트 생성...")
                    create_analysis_sga_sheet(wb, "Pivot",      "Analysis_SG&A",      target_period)
                    create_analysis_sga_sheet(wb, "Pivot(FCY)", "Analysis_SG&A(FCY)", target_period)

                    update_log("[Step 7] Sheet_total 생성...")
                    create_sheet_total(wb, "Pivot",      "Sheet_total PL",      target_period)
                    create_sheet_total(wb, "Pivot(FCY)", "Sheet_total PL(FCY)", target_period)

                    save_workbook(wb, config.OUTPUT_FILE)
                    update_log(captured.getvalue())
                    update_log(f"\n✅ 완료! → {config.OUTPUT_FILE}")

                config.TARGET_PERIOD = original_period

                st.success("✅ 분석 완료! '결과 다운로드' 탭에서 파일을 받으세요.")
                st.session_state["last_run_period"] = target_period

            except Exception:
                st.error("오류 발생:")
                st.code(traceback.format_exc())
                config.TARGET_PERIOD = original_period

# ===========================================================
# 탭 3: 결과 다운로드
# ===========================================================
with tab_result:
    st.subheader("📥 결과 파일 다운로드")

    output_file = config.OUTPUT_FILE

    if output_file.exists():
        import time
        mtime = output_file.stat().st_mtime
        mtime_str = pd.Timestamp(mtime, unit="s", tz="UTC").tz_convert("Asia/Seoul").strftime("%Y-%m-%d %H:%M")

        st.success(f"✓ 파일 생성됨: {output_file.name}  (최종 수정: {mtime_str})")

        with open(output_file, "rb") as f:
            file_bytes = f.read()

        st.download_button(
            label="⬇️  Excel 파일 다운로드",
            data=file_bytes,
            file_name=output_file.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

        # 생성된 시트 목록 미리보기
        st.divider()
        st.markdown("**포함된 시트**")
        try:
            import openpyxl
            wb_preview = openpyxl.load_workbook(output_file, read_only=True)
            cols = st.columns(2)
            for i, name in enumerate(wb_preview.sheetnames):
                cols[i % 2].markdown(f"- {name}")
            wb_preview.close()
        except Exception:
            pass
    else:
        st.info("아직 생성된 파일이 없습니다. '분석 실행' 탭에서 실행하세요.")
